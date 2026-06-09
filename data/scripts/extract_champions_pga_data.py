"""
extract_champions_pga_data.py
------------------------------
Reads Champions Network Garage_MC Attendance 2023.xlsx and extracts weekly
PGA (Physical Garage Attendance) by location.

Each monthly sheet has one row per location and one column per week date.
Location codes may be 5-char (pre-2023 era) or 6-char (current standard).
5-char codes are mapped to their current 6-char equivalents via CODE_MAP.

Outputs:
  data/outputs/champions_pga_data.json
    { locationCode (6-char), weekDate (YYYY-MM-DD), pga }
    Ready to feed into import-champions-pga.ts

  data/outputs/champions_unresolved_codes.json
    Records for old codes that could not be mapped to a current 6-char code.
    Each entry includes the old code, all week/pga records, and resolution
    instructions so a future operator can backfill without revisiting raw data.

Run:
  python3 data/scripts/extract_champions_pga_data.py
"""

import json
import os
from datetime import datetime
import openpyxl

EXCEL_FILE = os.path.join(
    os.path.dirname(__file__),
    "../Champions Data/Champions Network Garage_MC Attendance 2023.xlsx",
)
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "../outputs/champions_pga_data.json")
UNRESOLVED_FILE = os.path.join(os.path.dirname(__file__), "../outputs/champions_unresolved_codes.json")

# Sheets that are not weekly data (summaries, copies, special reports)
SKIP_SHEETS = {
    "Replenish", "Weekly Summary Report", "WHM BAPTISMSALVATIONS 23",
    "WHM SALNBAPTSM %", "Sheet4", "Copy of Sheet4", "Copy of 9th october",
    "Copy of  6th-31st August", "copy of 6th-31st August", "Copy of 6th-30th June",
}

# Full 5-char → 6-char location code mapping.
# Prefix-unambiguous mappings confirmed by code analysis.
# Name-resolved mappings confirmed by user + seed file lookup.
CODE_MAP: dict[str, str] = {
    # Prefix-unambiguous
    "WHBSG": "WHBSGA",
    "WHBSK": "WHBSKM",
    "WHBYG": "WHBYGR",
    "WHKIR": "WHKIRA",
    "WHKRN": "WHKRNY",
    "WHLYT": "WHLYTD",
    "WHMSK": "WHMSKA",
    "WHMTN": "WHMTND",
    "WHMTY": "WHMTYN",
    "WHSET": "WHSETA",
    # Resolved via user (location name) + seed file lookup
    "WHNLY": "WHNLYA",   # WH Naalya
    "WHSTA": "WHSETA",   # WH Seeta
    "WHNMV": "WHNMNV",   # WH Namanve
    "WHLTD": "WHLYTD",   # WH Lyantonde
    "WHKNG": "WHKUNG",   # WH Kungu
    "WHNKR": "WHNKWR",   # WH Nakwero
    # 6-char typo/transposition corrections (wrong code in source → correct code)
    "WHNKRW": "WHNKWR",  # transposition of WH Nakwero
    "WHNBGW": "WHNBSG",  # WH Nabusugwe (scrambled code)
    "WHBWGR": "WHBYGR",  # WH Bweyogerere (BW→BY transposition)
    # Unresolved — kept in champions_unresolved_codes.json:
    # WHKIT (5-113 PGA, Oct 2022–Jan 2023, no pastor name; likely WHKITI but unconfirmed)
    # WHKTSB (2 records; user suggested WH Kito-Kisooba but code pattern unconfirmed)
    # WHNLG, WHMLG (1 record each, negligible)
}

UNRESOLVED_CODES = {"WHKIT", "WHKTSB", "WHNLG", "WHMLG"}


def find_header_row(rows: list):
    """Return (row_index, {col_index: date_str}) for the first row with datetime values."""
    for i, row in enumerate(rows[:5]):
        date_cols = {
            j: cell.strftime("%Y-%m-%d")
            for j, cell in enumerate(row)
            if isinstance(cell, datetime)
        }
        if date_cols:
            return i, date_cols
    return None, None


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # (locationCode, weekDate) -> max pga seen (dedup across overlapping sheets)
    resolved: dict[tuple[str, str], int] = {}
    # oldCode -> { weekDate -> max pga }
    unresolved: dict[str, dict[str, int]] = {c: {} for c in UNRESOLVED_CODES}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=250, max_col=15, values_only=True))

        header_idx, date_cols = find_header_row(rows)
        if header_idx is None:
            continue

        for row in rows[header_idx + 1:]:
            raw_code = row[1] if len(row) > 1 else None
            if not isinstance(raw_code, str):
                continue
            raw_code = raw_code.strip()
            if not (raw_code.startswith("WH") and 4 <= len(raw_code) <= 6):
                continue

            for col_idx, date_str in date_cols.items():
                val = row[col_idx] if col_idx < len(row) else None
                if not isinstance(val, (int, float)) or val <= 0:
                    continue
                pga = int(val)

                if raw_code in UNRESOLVED_CODES:
                    unresolved[raw_code][date_str] = max(
                        unresolved[raw_code].get(date_str, 0), pga
                    )
                elif raw_code in CODE_MAP:
                    mapped = CODE_MAP[raw_code]
                    key = (mapped, date_str)
                    resolved[key] = max(resolved.get(key, 0), pga)
                elif len(raw_code) == 6:
                    # Current-era code with no correction needed
                    key = (raw_code, date_str)
                    resolved[key] = max(resolved.get(key, 0), pga)
                # else: unknown short code — silently skip

    wb.close()

    # Write resolved output
    records = [
        {"locationCode": code, "weekDate": date, "pga": pga}
        for (code, date), pga in sorted(resolved.items())
    ]
    location_codes = sorted(set(r["locationCode"] for r in records))
    week_dates = sorted(set(r["weekDate"] for r in records))

    output = {
        "description": "Champions Network weekly PGA data extracted from Garage_MC Attendance 2023.xlsx. "
                       "5-char legacy codes mapped to current 6-char codes. "
                       "Covers Feb 2022 – Feb 2024. Deduped: max value kept when a week appears in multiple sheets.",
        "recordCount": len(records),
        "locationCount": len(location_codes),
        "weekCount": len(week_dates),
        "dateRange": {"first": week_dates[0], "last": week_dates[-1]},
        "records": records,
    }
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)
    print(f"✅ Resolved: {len(records)} records | {len(location_codes)} locations | {len(week_dates)} weeks")
    print(f"   → {OUTPUT_FILE}")

    # Write unresolved output
    unresolved_entries = []
    for old_code, date_pga in unresolved.items():
        if not date_pga:
            continue
        sorted_records = [
            {"weekDate": d, "pga": p}
            for d, p in sorted(date_pga.items())
        ]
        dates = [r["weekDate"] for r in sorted_records]
        unresolved_entries.append({
            "oldCode": old_code,
            "recordCount": len(sorted_records),
            "dateRange": {"first": dates[0], "last": dates[-1]},
            "resolutionInstructions": (
                "Find the current 6-char WH**** location code for this location. "
                "Then re-run import-champions-pga.ts — it will skip records that already exist "
                "and add only new ones. No need to revisit the raw Excel file."
            ),
            "importTarget": "Sunday Service Report — pga field (ReportSubmissionData)",
            "records": sorted_records,
        })

    unresolved_output = {
        "description": "Champions Network PGA records for 5-char location codes that could not be "
                       "mapped to a current 6-char WH**** code. Resolve each oldCode to its current "
                       "code, add the mapping to CODE_MAP in extract_champions_pga_data.py, and re-run "
                       "the extraction. The importer is idempotent — already-imported records are skipped.",
        "locationCount": len(unresolved_entries),
        "totalRecords": sum(e["recordCount"] for e in unresolved_entries),
        "locations": unresolved_entries,
    }
    with open(UNRESOLVED_FILE, "w") as f:
        json.dump(unresolved_output, f, indent=2)
    unresolved_total = sum(e["recordCount"] for e in unresolved_entries)
    print(f"⚠️  Unresolved: {unresolved_total} records across {len(unresolved_entries)} codes → {UNRESOLVED_FILE}")


if __name__ == "__main__":
    main()
