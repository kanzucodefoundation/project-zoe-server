"""
extract_champions_oikos_2022.py
--------------------------------
Reads Champions Oikos 2022.xlsx and extracts weekly MC metrics per location.

Each sheet covers one week. Data rows have 5-char location codes in column A.
Codes are mapped to current 6-char equivalents via CODE_MAP (shared with PGA script).

Output:
  data/outputs/champions_oikos_2022.json
    Array of { locationCode, weekDate, mcs, mcm, mca, salvations, visitations }
    Ready to feed into import-champions-oikos-2022.ts

  data/outputs/champions_oikos_2022_unresolved.json
    Records for codes that could not be mapped.

Run:
  python3 data/scripts/extract_champions_oikos_2022.py
"""

import json
import os
import re
from datetime import datetime

import openpyxl

EXCEL_FILE = os.path.join(
    os.path.dirname(__file__),
    "../Champions Data/Champions Oikos 2022.xlsx",
)
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "../outputs/champions_oikos_2022.json")
UNRESOLVED_FILE = os.path.join(os.path.dirname(__file__), "../outputs/champions_oikos_2022_unresolved.json")

# Same CODE_MAP as extract_champions_pga_data.py
CODE_MAP: dict[str, str] = {
    "WHBSG": "WHBSGA",
    "WHBSK": "WHBSKM",
    "WHBYG": "WHBYGR",
    "WHKIR": "WHKIRA",
    "WHKRN": "WHKRNY",
    "WHLYT": "WHLYTD",
    "WHMSK": "WHMSKA",
    "WHMTN": "WHMTND",
    "WHMTY": "WHMTYN",
    "WHSET": "WHSETA",
    "WHNLY": "WHNLYA",
    "WHSTA": "WHSETA",
    "WHNMV": "WHNMNV",
    "WHLTD": "WHLYTD",
    "WHKNG": "WHKUNG",
    "WHNKR": "WHNKWR",
    "WHNKRW": "WHNKWR",
    "WHNBGW": "WHNBSG",
    "WHBWGR": "WHBYGR",
}

# Codes with no confirmed mapping
UNRESOLVED_CODES = {"WHKIT", "WHKTSB", "WHNLG", "WHMLG"}


def parse_sheet_date(name: str):
    """Return YYYY-MM-DD for a sheet name like '19th Feb 2023', else None."""
    name = re.sub(r"^Copy of ", "", name).strip()
    m = re.match(r"(\d+)(?:st|nd|rd|th)\s+(\w+)\s+(\d{4})", name)
    if not m:
        return None
    day, mon, yr = m.group(1), m.group(2), m.group(3)
    for fmt in ("%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(f"{int(day):02d} {mon} {yr}", fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def to_int(val):
    if val is None:
        return None
    if isinstance(val, str):
        # Skip formula strings
        if val.startswith("="):
            return None
        try:
            return int(float(val))
        except ValueError:
            return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # (locationCode, weekDate) -> record (dedup: last writer wins per sheet order)
    resolved: dict[tuple[str, str], dict] = {}
    unresolved: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        week_date = parse_sheet_date(sheet_name)
        if not week_date:
            continue  # Sheet3, Copy of Sheet1, etc.

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Data rows start at index 4 (row 5) through ~row 19
        for row in rows[4:25]:
            raw_code = row[0] if row[0] else None
            if not isinstance(raw_code, str):
                continue
            raw_code = raw_code.strip()
            if not (raw_code.startswith("WH") and 4 <= len(raw_code) <= 6):
                continue
            if "OVERALL" in raw_code.upper() or "NETWORK" in raw_code.upper():
                continue

            # Columns: A=location, B=leader, C=db_nos, D=mcs, E=mcm, F=mca,
            #          G=mca/db (formula), H=mca/mcm (formula), I=salvations, J=visitations
            mcs = to_int(row[3])
            mcm = to_int(row[4])
            mca = to_int(row[5])
            salvations = to_int(row[8])
            visitations = to_int(row[9])

            # Skip completely empty rows
            if all(v is None for v in [mcs, mcm, mca, salvations, visitations]):
                continue

            record = {
                "weekDate": week_date,
                "mcs": mcs,
                "mcm": mcm,
                "mca": mca,
                "salvations": salvations,
                "visitations": visitations,
            }

            if raw_code in UNRESOLVED_CODES:
                unresolved.setdefault(raw_code, []).append(record)
            elif raw_code in CODE_MAP:
                mapped = CODE_MAP[raw_code]
                resolved[(mapped, week_date)] = {**record, "locationCode": mapped}
            elif len(raw_code) == 6:
                resolved[(raw_code, week_date)] = {**record, "locationCode": raw_code}
            # else: unknown short code — skip

    records = sorted(resolved.values(), key=lambda r: (r["locationCode"], r["weekDate"]))
    location_codes = sorted(set(r["locationCode"] for r in records))
    week_dates = sorted(set(r["weekDate"] for r in records))

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(
            {
                "description": "Champions Oikos 2022-2023 weekly MC metrics per location",
                "recordCount": len(records),
                "locationCount": len(location_codes),
                "weekCount": len(week_dates),
                "dateRange": f"{week_dates[0]} – {week_dates[-1]}",
                "locationCodes": location_codes,
                "records": records,
            },
            f,
            indent=2,
        )

    print(f"✅ Resolved: {len(records)} records | {len(location_codes)} locations | {len(week_dates)} weeks")
    print(f"   Date range: {week_dates[0]} – {week_dates[-1]}")
    print(f"   Codes: {', '.join(location_codes)}")

    if unresolved:
        with open(UNRESOLVED_FILE, "w") as f:
            json.dump(unresolved, f, indent=2)
        print(f"⚠️  Unresolved codes: {', '.join(unresolved)} → {UNRESOLVED_FILE}")
    else:
        print("✅ No unresolved codes")


if __name__ == "__main__":
    main()
