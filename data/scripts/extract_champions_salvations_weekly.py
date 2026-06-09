"""
extract_champions_salvations_weekly.py
--------------------------------------
Reads weekly salvations (and baptisms where available) from:
  --year 2023  →  data/Champions Data/Champions SALVATIONS 2023.xlsx
  --year 2024  →  data/Champions Data/Champions Network 2024 SALVATIONS & BAPTISMS.xlsx

Each weekly sheet has one row per location with:
  col[WH Loc. col] = location name (partial, e.g. "Naalya") or WH-prefixed ("WHNaalya")
  col[Total col]   = total salvations
  col[Bapt. col]   = baptisms (present in 2024 and some 2023 sheets)

Sheet detection: Row 2 must contain "WH Loc." as a column header. Any sheet without it
(summary, monthly consolidated, or non-standard) is skipped.

Location name → 6-char code mapping: exact match after stripping "WH" prefix and lowercasing.
Unresolved names park in data/outputs/champions_salvations_weekly_{year}_unresolved.json.

Date extraction: parse from row 1 title "WHM SALVATION RESPONSES (DD MonthName YYYY)".
Start date is snapped to the nearest Sunday (max ±3 days).
Year falls back to the file's year when not explicit in the title.

Deduplication: (locationCode, weekDate) → max values when a location appears in multiple
sheets for the same week (common in 2024 due to overlapping/duplicate sheets).

Outputs:
  data/outputs/champions_salvations_weekly_{year}.json
  data/outputs/champions_salvations_weekly_{year}_unresolved.json

Run:
  python3 data/scripts/extract_champions_salvations_weekly.py --year 2023
  python3 data/scripts/extract_champions_salvations_weekly.py --year 2024
"""

from __future__ import annotations

import json
import os
import re
import sys
import argparse
from datetime import date, timedelta

import openpyxl

BASE = os.path.dirname(__file__)

FILES = {
    2023: os.path.join(BASE, "../Champions Data/Champions SALVATIONS 2023.xlsx"),
    2024: os.path.join(BASE, "../Champions Data/Champions Network 2024 SALVATIONS & BAPTISMS.xlsx"),
}

# Sheets to skip regardless of format (summaries, monthly consolidations)
SKIP_SHEETS = {
    # 2023
    "Form Responses 1", "2023 Summary",
    # 2024
    "2024 Summary", "FEB REPORT", "Copy of FEB REPORT", "January Summary", "JAN 2024",
}

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Location name (lowercase, WH-prefix stripped) → 6-char code
# Built from whm-group-tree.seed.ts (all 226 seeded locations)
NAME_TO_CODE: dict[str, str] = {
    "abayita": "WHABYT",
    "africa online": "WHAFOL",
    "apac": "WHAONL",
    "arua": "WHARUA",
    "asia online": "WHASOL",
    "australia": "WHAUST",
    "bajjo": "WHBAJO",
    "bermuda": "WHBMDA",
    "biharwe": "WHBHRW",
    "bondo": "WHBNDO",
    "budaka": "WHBDKA",
    "buddo": "WHBUDO",
    "budondo": "WHBDND",
    "bugembe": "WHBGMB",
    "bugolobi": "WHBGLB",
    "bukasa": "WHBKSA",
    "bukerere": "WHBKRR",
    "bukoto": "WHBKTO",
    "bulindo": "WHBKYA",
    "buloba": "WHBULB",
    "bulondo": "WHBLND",
    "bunamwaya": "WHBNMY",
    "busabala": "WHBSBL",
    "busega": "WHBSGA",
    "bushenyi": "WHBSHY",
    "busiika": "WHBWBJ",
    "busukuma": "WHBSKM",
    "buwaya": "WHBWYA",
    "buwenge": "WHBWNG",
    "buziga": "WHBZGA",
    "bwebajja": "WHBWGA",
    "bwerenga": "WHBWGA",
    "bweyogerere": "WHBYGR",
    "canada": "WHCNDA",
    "dar es salaam": "WHDNTN",
    "downtown": "WHDWTN",
    "entebbe": "WHENTB",
    "entebbe central": "WHEBCT",
    "europe online": "WHEONL",
    "fortportal": "WHFTPT",
    "fort portal": "WHFTPT",
    "garuga": "WHGRUG",
    "gayaza": "WHGYZA",
    "geneva": "WHGNVA",
    "germany": "WHGMNY",
    "ggaba road": "WHGBRD",
    "gomba": "WHGOMB",
    "gulu": "WHGULU",
    "hoima": "WHHOMA",
    "ibanda": "WHIBND",
    "idudi": "WHIDDI",
    "iganga": "WHIGNG",
    "iganga cms": "WHICMS",
    "ishaka": "WHISHK",
    "jinja": "WHJNJA",
    "joggo": "WHJOGO",
    "juja": "WHJUJA",
    "kabale": "WHKBLE",
    "kabembe": "WHKBMB",
    "kabubu": "WHKBUB",
    "kaihura": "WHKGWA",
    "kakira": "WHKKRA",
    "kakiri": "WHKKRI",
    "kakoba": "WHKKBA",
    "kalagi": "WHKLGI",
    "kaliro": "WHKLRO",
    "kamuli": "WHKMLI",
    "kasanje": "WHKSNJ",
    "kasenge": "WHKSKS",
    "kasengejje": "WHKSGJ",
    "kasese": "WHKSSE",
    "kasubi": "WHKSBI",
    "katende": "WHKTND",
    "kavule": "WHKVLE",
    "kawanda": "WHKWND",
    "kawempe": "WHKWMP",
    "kayunga": "WHKYJO",
    "kayunga - wakiso": "WHKYGW",
    "kenya online": "WHKEOL",
    "kericho": "WHKMWK",
    "kiamu": "WHKMBU",
    "kibibi": "WHKBBI",
    "kibinge": "WHKBNG",
    "kibuli": "WHKBLI",
    "kibuye": "WHKBYE",
    "kigali": "WHKGLI",
    "kigumba": "WHKGMB",
    "kigungu": "WHKGNG",
    "kijabijjo": "WHKJBJ",
    "kilemezi": "WHKLMZ",
    "kimwanyi": "WHKMNY",
    "kira": "WHKIRA",
    "kirinya": "WHKRNY",
    "kisaasi": "WHKSAS",
    "kisoga": "WHKISG",
    "kitende": "WHKITD",
    "kitengela": "WHKTGL",
    "kiti": "WHKITI",
    "kitikifumba": "WHKTFB",
    "kito-kisooba": "WHKITO",
    "kito kisooba": "WHKITO",
    "kitokilsooba": "WHKITO",
    "kitukutwe": "WHKTKT",
    "kiwanga": "WHKWGA",
    "kiwenda": "WHKWED",
    "koboko": "WHKBKO",
    "krefeld": "WHKFLD",
    "kulambiro": "WHKLBR",
    "kungu": "WHKUNG",
    "kyanja": "WHKYNJ",
    "kyebando": "WHKYBD",
    "kyegegwa": "WHKYGG",
    "kyengera": "WHKYGR",
    "kyenjojo": "WHKYJJ",
    "kyetume": "WHKYTM",
    "leeds": "WHLEDS",
    "lira": "WHLIRA",
    "lugazi": "WHLUGZ",
    "lugoba": "WHLGBA",
    "lukaya": "WHLKYA",
    "lungujja": "WHLGJA",
    "lusaka": "WHLUSK",
    "lusanja": "WHLSNJ",
    "lusaze": "WHLUSZ",
    "luton": "WHLUTN",
    "luuka": "WHLUKA",
    "luwero": "WHLWRO",
    "luzira": "WHLZRA",
    "lyantonde": "WHLYTD",
    "mafubira": "WHMFBR",
    "magamaga": "WHMGMG",
    "maganjo": "WHMGJO",
    "magere": "WHMGRE",
    "makerere": "WHMKRR",
    "makindye": "WHMKND",
    "masaka": "WHMSKA",
    "masanafu": "WHMSNF",
    "masindi": "WHMSND",
    "masulita": "WHMSLT",
    "matugga": "WHMTUG",
    "mawanda road": "WHMWRD",
    "maya": "WHMASK",
    "mbale": "WHMBLE",
    "mbarara": "WHMBRA",
    "mbiko": "WHMBKO",
    "mengo": "WHMNGO",
    "misindye": "WHMSDY",
    "mityana": "WHMTWE",
    "mombasa": "WHMBSA",
    "mpererwe": "WHMPRW",
    "mpigi": "WHMPGI",
    "mubende": "WHMBND",
    "mukono": "WHMKNO",
    "mukono central": "WHMKNC",
    "mukono city": "WHMCTY",
    "mutai": "WHMTAI",
    "mutundwe": "WHMTND",
    "mutungo-biina": "WHMTGB",
    "muwafu": "WHMWFU",
    "muyenga": "WHMYNG",
    "naalya": "WHNLYA",
    "nabusugwe": "WHNBSG",
    "nairobi": "WHNRBI",
    "najjera": "WHNAJJ",
    "nakasajja": "WHNKSJ",
    "nakaseke": "WHNKSK",
    "nakawa": "WHNKWA",
    "nakawuka": "WHNKWK",
    "nakifuma": "WHNKFM",
    "nakulabye": "WHNKLB",
    "nakwero": "WHNKWR",
    "nalumunye": "WHNLMY",
    "namafresh": "WHNMFR",
    "namagera": "WHNMGR",
    "namanve": "WHNMNV",
    "namataba": "WHNMTB",
    "namilyango": "WHNLYG",
    "nampunge": "WHNMPG",
    "namulesa": "WHNLSA",
    "namulonge": "WHNMLG",
    "namuwongo": "WHNMWG",
    "namwendwa": "WHNMVD",
    "nangabo": "WHNGBO",
    "nansana": "WHNSNA",
    "nansana gganda": "WHNGND",
    "nawampanda": "WHNWPD",
    "ndejje": "WHNDJE",
    "ndejje-bombo": "WHNDJB",
    "nebbi": "WHNEBB",
    "njeru": "WHNJRU",
    "nkonkojeru": "WHNKNJ",
    "nkumba": "WHNKMB",
    "north wales": "WHNWLS",
    "nsaggu": "WHNSGU",
    "nsambwe": "WHNSBW",
    "nsambya": "WHNSBY",
    "nsangi": "WHNSGI",
    "nsasa": "WHNSSA",
    "ntawo": "WHNTND",
    "ntinda": "WHNTDA",
    "ntungamo": "WHNTGM",
    "nyendo": "WHNYDO",
    "ongata rongai": "WHONGR",
    "paidha": "WHPDHA",
    "pakwach": "WHPKCH",
    "pallisa": "WHPLSA",
    "rubaga": "WHRBGA",
    "ruiru": "WHRUIR",
    "salama road": "WHSLRD",
    "seeta": "WHSETA",
    "seguku": "WHSEGK",
    "sentema": "WHSNTM",
    "sonde": "WHSOND",
    "soroti": "WHSROT",
    "texas": "WHTEXS",
    "thika": "WHTHKA",
    "tororo": "WHTRRO",
    "tula": "WHTULA",
    "united kingdom": "WHUKDM",
    "united states of america": "WHUSOA",
    "usa": "WHUSOA",
    "wairaka": "WHWRKA",
    "wakiso": "WHWKSO",
    "wandi": "WHWNDI",
    "wattuba": "WHWTBA",
    "wigan": "WHWGAN",
    "wobulenzi": "WHWBLZ",
    "yesu amala": "WHNYSA",
    "zirobwe": "WHZRBW",
    "kajjansi": "WHKJNS",
    # Common truncated/variant spellings found in source data
    "bwebaja": "WHBWGA",
    "nansana main": "WHNSNA",
    "kibunge": "WHKBNG",
    "kito": "WHKITO",
    # Typos / alternate spellings seen in weekly salvations sheets
    "jjogo": "WHJOGO",
    "salaama road": "WHSLRD",
    "bugoloobi": "WHBGLB",
    "kabubbu": "WHKBUB",
    "kawompe": "WHKWMP",
    "kawempe ": "WHKWMP",
    "kasokoso": "WHKSAS",   # Kasokoso is in the Kisaasi area
    "njeru rd re": "WHNJRU",
    "njeru rd": "WHNJRU",
    "kisasi": "WHKSAS",
    "entebbe ctrl": "WHEBCT",
    "entebbe central": "WHEBCT",
    "mukono ctrl": "WHMKNC",
    "uk": "WHUKDM",
    "kayunga wakiso": "WHKYGW",
    "kayunga-wakiso": "WHKYGW",
}


def normalize_name(raw: str) -> str:
    """Strip WH/wh prefix and leading/trailing whitespace, lowercase."""
    s = raw.strip()
    if s.upper().startswith("WH"):
        s = s[2:].lstrip()
    return s.lower().strip()


def lookup_code(raw_name: str) -> str | None:
    s = raw_name.strip()

    # Direct 6-char code (e.g. "WHBDKA") — pass through
    if re.match(r"^WH[A-Z0-9]{4}$", s.upper()):
        return s.upper()

    # 4-char suffix code without WH prefix (e.g. "BDKA" from swapped-column sheets)
    if re.match(r"^[A-Z0-9]{4}$", s.upper()):
        return "WH" + s.upper()

    # Name-based lookup
    norm = normalize_name(s)
    code = NAME_TO_CODE.get(norm)
    if code:
        return code
    norm2 = norm.rstrip(" .")
    return NAME_TO_CODE.get(norm2)


def parse_week_start_date(title: str, default_year: int) -> str | None:
    """
    Parse the start-of-week Sunday date from a title like:
    "WHM SALVATION RESPONSES (28th May-3rd June 2023)"
    "WHM SALVATION RESPONSES (1st-6th Jan 2024)"
    "WHM SALVATION RESPONSES (28th-4th June 2023)"
    Returns YYYY-MM-DD string, snapped to nearest Sunday (±3 days).
    """
    # Extract parenthetical
    m = re.search(r"\((.+?)\)", title)
    content = m.group(1).strip() if m else title.strip()

    # Extract year
    year_m = re.search(r"\b(20\d{2})\b", content)
    year = int(year_m.group(1)) if year_m else default_year

    # Extract month names
    month_matches = re.findall(
        r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\b", content, re.I
    )
    # Extract day numbers
    day_matches = re.findall(r"\b(\d{1,2})(?:st|nd|rd|th)?\b", content)
    days = [int(d) for d in day_matches]

    if not days or not month_matches:
        return None

    first_day = days[0]

    if len(month_matches) >= 2:
        # Range crosses months: "28th May-3rd June" — first month named explicitly
        first_month = MONTH_MAP.get(month_matches[0].lower()[:3])
        if not first_month:
            return None
    else:
        end_month = MONTH_MAP.get(month_matches[0].lower()[:3])
        if not end_month:
            return None
        if len(days) >= 2:
            last_day = days[-1]
            if first_day > last_day:
                # Start is in the previous month: "28th-4th June" → May 28
                first_month = end_month - 1 if end_month > 1 else 12
                if end_month == 1:
                    year -= 1
            else:
                first_month = end_month
        else:
            first_month = end_month

    try:
        start = date(year, first_month, first_day)
    except ValueError:
        return None

    # Snap to nearest Sunday (weekday 6)
    for delta in range(-3, 4):
        candidate = start + timedelta(days=delta)
        if candidate.weekday() == 6:
            return candidate.strftime("%Y-%m-%d")

    # No Sunday within ±3 days — use start date as-is
    return start.strftime("%Y-%m-%d")


def detect_columns(header_row2: list) -> tuple[int | None, int | None, int | None]:
    """
    Return (loc_col, total_col, bapt_col) from the header row.
    Searches for 'WH Loc.' (or 'Location'), 'Total', 'Bapt.' labels.
    """
    loc_col = total_col = bapt_col = None
    for i, cell in enumerate(header_row2):
        if not isinstance(cell, str):
            continue
        s = cell.strip().lower()
        if s in ("wh loc.", "wh loc", "location") and loc_col is None:
            loc_col = i
        elif s == "total" and total_col is None:
            total_col = i
        elif s.startswith("bapt") and bapt_col is None:
            bapt_col = i
    return loc_col, total_col, bapt_col


def process_sheet(ws, default_year: int) -> tuple[str | None, list[tuple[str, int, int | None]]]:
    """
    Process one weekly sheet. Returns (weekDate, [(rawName, salvations, baptisms|None)]).
    Returns (None, []) if the sheet is not a recognised weekly format.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=12, values_only=True))
    if len(rows) < 3:
        return None, []

    # Row 1 title
    title_row = rows[0]
    title = str(title_row[0]).strip() if title_row[0] else ""

    # Row 2 header
    header_row = rows[1]
    loc_col, total_col, bapt_col = detect_columns(header_row)
    if loc_col is None or total_col is None:
        return None, []

    week_date = parse_week_start_date(title, default_year)
    if not week_date:
        return None, []

    results = []
    for row in rows[2:]:
        # Data rows have a numeric No. in col 0
        no_val = row[0] if len(row) > 0 else None
        if not isinstance(no_val, (int, float)) or isinstance(no_val, bool):
            continue

        raw_name = row[loc_col] if len(row) > loc_col else None
        if not isinstance(raw_name, str) or not raw_name.strip():
            continue

        total_val = row[total_col] if len(row) > total_col else None
        salvs = int(total_val) if isinstance(total_val, (int, float)) and not isinstance(total_val, bool) else 0
        if salvs < 0:
            salvs = 0

        bapts = None
        if bapt_col is not None:
            bapt_val = row[bapt_col] if len(row) > bapt_col else None
            if isinstance(bapt_val, (int, float)) and not isinstance(bapt_val, bool):
                bapts = int(bapt_val)
                if bapts < 0:
                    bapts = 0

        results.append((raw_name.strip(), salvs, bapts))

    return week_date, results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, choices=[2023, 2024], required=True)
    args = parser.parse_args()

    year = args.year
    excel_file = FILES[year]
    output_file = os.path.join(BASE, f"../outputs/champions_salvations_weekly_{year}.json")
    unresolved_file = os.path.join(BASE, f"../outputs/champions_salvations_weekly_{year}_unresolved.json")

    if not os.path.exists(excel_file):
        print(f"❌ File not found: {excel_file}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # (code, weekDate) → (max_salvations, max_baptisms | None)
    resolved: dict[tuple[str, str], tuple[int, int | None]] = {}
    # rawName → set of weekDates where unresolved
    unresolved: dict[str, set] = {}

    sheets_processed = 0
    sheets_skipped = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            sheets_skipped += 1
            continue
        if sheet_name.lower().startswith("copy of"):
            sheets_skipped += 1
            continue

        ws = wb[sheet_name]
        week_date, rows = process_sheet(ws, year)
        if not week_date or not rows:
            sheets_skipped += 1
            continue

        sheets_processed += 1
        for raw_name, salvs, bapts in rows:
            code = lookup_code(raw_name)
            if not code:
                unresolved.setdefault(raw_name, set()).add(week_date)
                continue
            key = (code, week_date)
            prev_s, prev_b = resolved.get(key, (0, None))
            new_b = max(prev_b or 0, bapts or 0) if (prev_b is not None or bapts is not None) else None
            resolved[key] = (max(prev_s, salvs), new_b)

    wb.close()

    records = [
        {
            "locationCode": code,
            "weekDate": week_date,
            "salvations": salvs,
            **({"baptisms": bapts} if bapts is not None else {}),
        }
        for (code, week_date), (salvs, bapts) in sorted(resolved.items())
    ]

    location_codes = sorted(set(r["locationCode"] for r in records))
    week_dates = sorted(set(r["weekDate"] for r in records))

    output = {
        "description": (
            f"Champions Network weekly salvations data for {year}. "
            f"Location names fuzzy-matched to 6-char codes. "
            f"Deduped: max value kept when a location appears in multiple sheets for the same week."
        ),
        "year": year,
        "recordCount": len(records),
        "locationCount": len(location_codes),
        "weekCount": len(week_dates),
        "sheetsProcessed": sheets_processed,
        "sheetsSkipped": sheets_skipped,
        "dateRange": {"first": week_dates[0] if week_dates else None, "last": week_dates[-1] if week_dates else None},
        "records": records,
    }
    with open(output_file, "w") as f:
        json.dump(output, f, indent=2)
    print(f"✅ {year}: {len(records)} records | {len(location_codes)} locations | {len(week_dates)} weeks")
    print(f"   Sheets processed: {sheets_processed}, skipped: {sheets_skipped}")
    print(f"   → {output_file}")

    unresolved_entries = [
        {
            "rawName": name,
            "weekCount": len(dates),
            "weekDates": sorted(dates),
            "resolutionInstructions": (
                "Find this location in the group tree seed (src/seed/whm/whm-group-tree.seed.ts) "
                "and add the mapping to NAME_TO_CODE in this script. Re-run to re-extract."
            ),
        }
        for name, dates in sorted(unresolved.items())
    ]
    unresolved_output = {
        "description": f"Weekly salvations {year}: location names that could not be mapped to a 6-char code.",
        "locationCount": len(unresolved_entries),
        "locations": unresolved_entries,
    }
    with open(unresolved_file, "w") as f:
        json.dump(unresolved_output, f, indent=2)
    if unresolved_entries:
        print(f"⚠️  Unresolved: {len(unresolved_entries)} location names → {unresolved_file}")
    else:
        print(f"✅ No unresolved location names.")


if __name__ == "__main__":
    main()
