"""
extract_mca_data.py
-------------------
Reads WHM MC ATTENDANCE 2023-2024 (13).xlsx and extracts weekly MCA
(Missional Community Attendance) by location code.

Output: data/outputs/mca_weekly_data.json
  { locationCode, weekDate (YYYY-MM-DD), mca }

These records will be imported as the `mca` field on existing Sunday Service
Report submissions (or as new submissions where no PGA record exists).
"""

import json
import re
import os
from datetime import datetime
import openpyxl

EXCEL_FILE = os.path.join(
    os.path.dirname(__file__),
    "../More Data/WHM MC ATTENDANCE 2023 -2024 (13).xlsx",
)
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "../outputs/mca_weekly_data.json")

# Sheets to extract, in preference order. Later sheets for the same
# (code, date) overwrite earlier ones — "New Initials" sheets are more
# complete than the plain "January " draft.
TARGET_SHEETS = [
    "January ",                    # draft — overwritten by New Initials below
    "January New Initials 2023",
    "February New Initials 2023",
    "March New Initials 2023",
    "April New Initials 2023",
    "May New Initials 2023",
    # June 2023 absent from source file
    "July New Initials 2023 2",
    "Aug New Initials 2023 ",
    "Sept New Initials 2023  ",
    "October New Initials 2023  ",
    "November New Initials 2023",
    # December 2023 absent from source file
    "January 2024",
    "February 2024",
    "March 2024 ",
    "April 2024",
]

LOCATION_CODE_RE = re.compile(r"^WH[A-Z]{4}\s*$", re.IGNORECASE)


def is_location_code(value):
    return isinstance(value, str) and bool(LOCATION_CODE_RE.match(value.strip()))


def find_header_row(ws):
    """Return (header_row_index, {col_index: date}) for the first row
    that contains datetime objects in any column past index 2."""
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        date_cols = {
            j: v
            for j, v in enumerate(row)
            if isinstance(v, datetime)
        }
        if date_cols:
            return i, date_cols
    return None, {}


def extract_sheet(ws):
    """Return list of {locationCode, weekDate, mca} from one sheet."""
    header_row_idx, date_cols = find_header_row(ws)
    if not date_cols:
        return []

    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue

        # Location code is whichever column (0-3) looks like WH[A-Z]{4}
        code = None
        for j in range(min(4, len(row))):
            if is_location_code(str(row[j] or "")):
                code = row[j].strip().upper()
                break
        if not code:
            continue

        for col_idx, dt in date_cols.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "" or val == "-":
                continue
            try:
                mca = int(float(str(val)))
            except (ValueError, TypeError):
                continue
            if mca <= 0:
                continue

            records.append({
                "locationCode": code,
                "weekDate": dt.strftime("%Y-%m-%d"),
                "mca": mca,
            })

    return records


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    available = set(wb.sheetnames)

    # keyed by (locationCode, weekDate) — later sheets overwrite earlier
    deduped: dict[tuple, dict] = {}
    sheet_stats = []

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in available:
            print(f"  [skip] Sheet not found: '{sheet_name}'")
            continue
        ws = wb[sheet_name]
        records = extract_sheet(ws)
        for rec in records:
            deduped[(rec["locationCode"], rec["weekDate"])] = rec
        sheet_stats.append((sheet_name.strip(), len(records)))
        print(f"  [ok]   {sheet_name.strip()}: {len(records)} records")

    records = sorted(deduped.values(), key=lambda r: (r["weekDate"], r["locationCode"]))

    dates = sorted(set(r["weekDate"] for r in records))
    locations = sorted(set(r["locationCode"] for r in records))

    print(f"\nTotal: {len(records)} records | {len(locations)} locations | {len(dates)} weeks")
    print(f"Date range: {dates[0]} → {dates[-1]}")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    output = {
        "recordCount": len(records),
        "locationCount": len(locations),
        "weekCount": len(dates),
        "dateRange": {"from": dates[0], "to": dates[-1]},
        "records": records,
    }
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Written to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
