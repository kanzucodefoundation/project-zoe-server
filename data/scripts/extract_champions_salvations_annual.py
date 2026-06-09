"""
extract_champions_salvations_annual.py
--------------------------------------
Reads the WHM BAPTISMSALVATIONS 23 sheet from:
  data/Champions Data/Champions Network Garage_MC Attendance 2023.xlsx

Two side-by-side tables (cols 0-4 and cols 6-10).
Each table has category sections (CHAMPIONS, KINGDOM ADVANCER, etc.) followed by data rows.
Data rows have a numeric No. in col 0/6 and a WH* location code in col 1/7.

Deduplication: when a code appears in both left and right tables, take the max value.
reportingPeriod: 2023-12-31 (cumulative 2023 totals).

CODE_MAP: WHKUNGU (7-char) → WHKUNG (confirmed typo, extra U suffix).
Unresolved: WHBLB, WHBMB, WHKRO, WHNFM — 5-char codes with no confirmed 6-char mapping.
WHSLRD rows have '-' values — skipped (no numeric data).

Outputs:
  data/outputs/champions_salvations_annual_2023.json
  data/outputs/champions_salvations_annual_unresolved.json

Run:
  python3 data/scripts/extract_champions_salvations_annual.py
"""

from __future__ import annotations

import json
import os

import openpyxl

BASE = os.path.dirname(__file__)
EXCEL_FILE = os.path.join(BASE, "../Champions Data/Champions Network Garage_MC Attendance 2023.xlsx")
OUTPUT_FILE = os.path.join(BASE, "../outputs/champions_salvations_annual_2023.json")
UNRESOLVED_FILE = os.path.join(BASE, "../outputs/champions_salvations_annual_unresolved.json")

SHEET_NAME = "WHM BAPTISMSALVATIONS 23"

CODE_MAP: dict[str, str] = {
    "WHKUNGU": "WHKUNG",  # 7-char typo — extra U suffix, confirmed WH Kungu
}

UNRESOLVED_CODES = {"WHBLB", "WHBMB", "WHKRO", "WHNFM"}

SKIP_VALUES = {"-", "N/A", ""}


def is_numeric(val) -> bool:
    return isinstance(val, (int, float)) and not isinstance(val, bool)


def parse_int(val) -> int | None:
    if isinstance(val, float):
        return int(val)
    if isinstance(val, int):
        return val
    return None


def extract_table(rows: list, code_col: int, salvs_col: int, bapt_col: int) -> dict[str, tuple[int, int]]:
    """Extract (salvations, baptisms) per code from one table column set."""
    result: dict[str, tuple[int, int]] = {}
    for row in rows:
        no_val = row[0] if len(row) > 0 else None
        if not is_numeric(no_val):
            continue

        raw_code = row[code_col] if len(row) > code_col else None
        if not isinstance(raw_code, str):
            continue
        code = raw_code.strip()
        if not code.startswith("WH") or len(code) < 4:
            continue

        salvs_raw = row[salvs_col] if len(row) > salvs_col else None
        bapt_raw = row[bapt_col] if len(row) > bapt_col else None

        # Skip rows with placeholder text values
        if str(salvs_raw).strip() in SKIP_VALUES and str(bapt_raw).strip() in SKIP_VALUES:
            continue

        salvs = parse_int(salvs_raw) or 0
        bapt = parse_int(bapt_raw) or 0

        # Merge: take max per field
        prev_s, prev_b = result.get(code, (0, 0))
        result[code] = (max(prev_s, salvs), max(prev_b, bapt))

    return result


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found in {EXCEL_FILE}")
        return
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=12, values_only=True))
    wb.close()

    # Left table: No=col0, code=col1, salvs=col2, bapt=col3
    left = extract_table(rows, code_col=1, salvs_col=2, bapt_col=3)
    # Right table: No=col6, code=col7, salvs=col8, bapt=col9
    right = extract_table(rows, code_col=7, salvs_col=8, bapt_col=9)

    # Merge both tables; take max where both have data for same code
    combined: dict[str, tuple[int, int]] = {}
    for code, vals in {**left, **right}.items():
        prev = combined.get(code, (0, 0))
        combined[code] = (max(prev[0], vals[0]), max(prev[1], vals[1]))

    resolved: list[dict] = []
    unresolved_map: dict[str, dict] = {}

    for code, (salvs, bapt) in sorted(combined.items()):
        if code in UNRESOLVED_CODES:
            unresolved_map[code] = {"salvations": salvs, "baptisms": bapt}
            continue
        if code in CODE_MAP:
            code = CODE_MAP[code]
        if len(code) != 6:
            # Non-standard length not in map — park as unresolved
            unresolved_map[code] = {"salvations": salvs, "baptisms": bapt}
            continue
        resolved.append({"locationCode": code, "salvations": salvs, "baptisms": bapt})

    output = {
        "description": (
            "Annual 2023 salvations and baptisms per WHM location. "
            "Extracted from WHM BAPTISMSALVATIONS 23 sheet. "
            "Both left and right tables merged; max value kept when a location appears in both. "
            "reportingPeriod for all records: 2023-12-31."
        ),
        "reportingPeriod": "2023-12-31",
        "recordCount": len(resolved),
        "locationCount": len(resolved),
        "records": resolved,
    }
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)
    print(f"✅ Resolved: {len(resolved)} locations → {OUTPUT_FILE}")

    unresolved_entries = [
        {
            "code": code,
            "salvations": vals["salvations"],
            "baptisms": vals["baptisms"],
            "resolutionInstructions": (
                "Find the current 6-char WH**** code for this location. "
                "Add it to CODE_MAP in this script and re-run. "
                "The importer is idempotent — already-imported records are skipped."
            ),
            "importTarget": "Annual Salvations & Baptisms report — salvations and baptisms fields",
        }
        for code, vals in sorted(unresolved_map.items())
    ]
    unresolved_output = {
        "description": (
            "Annual 2023 salvations records for codes that could not be mapped to a "
            "current 6-char WH**** code. Resolve each code, add to CODE_MAP, and re-run."
        ),
        "locationCount": len(unresolved_entries),
        "locations": unresolved_entries,
    }
    with open(UNRESOLVED_FILE, "w") as f:
        json.dump(unresolved_output, f, indent=2)
    print(f"⚠️  Unresolved: {len(unresolved_entries)} codes → {UNRESOLVED_FILE}")


if __name__ == "__main__":
    main()
