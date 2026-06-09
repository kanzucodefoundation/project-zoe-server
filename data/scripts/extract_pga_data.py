"""
extract_pga_data.py
-------------------
Reads every PGA source across the data/ folder and outputs a single
deduplicated JSON file: data/outputs/pga_weekly_data.json

Record shape:
  { "locationCode": "WHNLYA", "weekDate": "2024-01-07", "pga": 2182 }

Source priority (higher overwrites lower for same location+date):
  3 - individual slot sheet  (actual weekly, most detailed)
  2 - DIA PDF                (actual weekly)
  1 - continuous / 5-wk xlsx (actual weekly column)
  0 - 4-week average PDF     (column values are still per-week, not averaged)

Run from the server/ directory:
  python3 data/scripts/extract_pga_data.py
"""

from __future__ import annotations

import glob
import json
import os
import re
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]          # data/
OUTPUT_DIR = ROOT / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "pga_weekly_data.json"

# Known WHM location codes — used as whitelist to suppress noise.
# Loaded from worship_harvest_categorization_data.json at runtime.
KNOWN_CODES: set[str] = set()

MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalise_code(raw: Any) -> str | None:
    """Return a clean 6-char WH[A-Z]{4} code, or None if unrecognisable.
    Uses KNOWN_CODES as a whitelist when populated."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    # All real WHM codes are exactly WH + 4 uppercase letters = 6 chars total
    if not re.match(r'^WH[A-Z]{4}$', s):
        return None
    if KNOWN_CODES and s not in KNOWN_CODES:
        return None
    return s


def to_iso(d: Any) -> str | None:
    """Convert an openpyxl date/datetime/date object to YYYY-MM-DD string."""
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y-%m-%d') if hasattr(d, 'year') else None
    return None


def parse_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        s = str(v).replace(',', '').strip()
        f = float(s)
        return int(round(f)) if f > 0 else None
    except (ValueError, TypeError):
        return None


def parse_date_from_string(text: str) -> date | None:
    """Parse 'May 10, 2026' / '3-May' / '10th May 2026' etc."""
    text = text.strip()
    # Full date: "10th May 2026" or "10 May 2026" or "May 10, 2026"
    m = re.search(
        r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)[,\s]+(\d{4})',
        text, re.IGNORECASE
    )
    if m:
        day, mon, yr = int(m.group(1)), m.group(2).lower()[:3], int(m.group(3))
        if mon in MONTH_MAP:
            return date(yr, MONTH_MAP[mon], day)
    m2 = re.search(
        r'([A-Za-z]+)\s+(\d{1,2})[,\s]+(\d{4})',
        text, re.IGNORECASE
    )
    if m2:
        mon, day, yr = m2.group(1).lower()[:3], int(m2.group(2)), int(m2.group(3))
        if mon in MONTH_MAP:
            return date(yr, MONTH_MAP[mon], day)
    return None


def parse_short_date(text: str, year: int) -> date | None:
    """Parse '3-May', '12-Apr', '19 Apr' within a known year."""
    text = text.strip()
    m = re.match(r'(\d{1,2})[-\s]([A-Za-z]+)', text)
    if m:
        day, mon = int(m.group(1)), m.group(2).lower()[:3]
        if mon in MONTH_MAP:
            return date(year, MONTH_MAP[mon], day)
    return None


# ---------------------------------------------------------------------------
# Source readers
# ---------------------------------------------------------------------------

Records = dict[tuple[str, str], dict]   # (code, iso_date) -> {pga, priority}


def add(records: Records, code: str, iso_date: str, pga: int, priority: int) -> None:
    key = (code, iso_date)
    if key not in records or records[key]['priority'] < priority:
        records[key] = {'pga': pga, 'priority': priority}


# ── Excel: continuous / 5-week average sheets ───────────────────────────────

def read_continuous_excel(records: Records, path: str) -> int:
    """Parse any sheet whose header row contains datetime columns after the location col."""
    count = 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if any(x in sh.lower() for x in ['english', 'cathedral', 'hectic', 'rough', 'copy']):
                continue
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue
            # find header row: must have datetime cols after col index 1
            header_row = None
            header_idx = None
            for i, row in enumerate(rows[:4]):
                date_cols = [(j, to_iso(c)) for j, c in enumerate(row) if to_iso(c)]
                if len(date_cols) >= 2:
                    header_row = row
                    header_idx = i
                    break
            if not header_row:
                continue
            date_cols = [(j, to_iso(c)) for j, c in enumerate(header_row) if to_iso(c)]
            for row in rows[header_idx + 1:]:
                code = normalise_code(row[1] if len(row) > 1 else None)
                if not code:
                    continue
                for col_idx, iso_date in date_cols:
                    if col_idx >= len(row):
                        continue
                    pga = parse_int(row[col_idx])
                    if pga and pga > 0:
                        add(records, code, iso_date, pga, priority=1)
                        count += 1
    except Exception as e:
        print(f'  [warn] {os.path.basename(path)}: {e}')
    return count


# ── Excel: individual slot sheets ────────────────────────────────────────────

SLOT_COL_MAP = {
    # Normalise whatever header names appear to our canonical field names
    '1sv': '1Sv', 'firstservice': '1Sv', '9:00': '1Sv', '9:00 am': '1Sv',
    '2sv': '2Sv', 'secondservice': '2Sv', '11:30': '2Sv', '11:30 am': '2Sv',
    'yxp': 'YXP', 'teens': 'YXP',
    'kids': 'kids', 'kid': 'kids',
    'local': 'local', 'local l.': 'local', 'loc': 'local',
    'hc1': 'hc1', 'hostingcenter1': 'hc1',
    'hc2': 'hc2', 'hostingcenter2': 'hc2',
    'hc3': 'hc3', 'hostingcenter3': 'hc3',
    'alc': 'alc',
    'ftg': 'ftg',
}


def read_slot_excel(records: Records, path: str, iso_date: str) -> int:
    """Read a single-week slot sheet and store as priority-3 record."""
    count = 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if any(x in sh.lower() for x in
                   ['english', 'average', 'continuous', 'copy', 'epga', 'cathedral', 'hectic']):
                continue
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
            # Find header row with slot columns
            header_row = None
            header_idx = None
            for i, row in enumerate(rows[:4]):
                cols = [str(c).strip().lower() if c else '' for c in row]
                if any(k in cols for k in ['1sv', '2sv', '9:00', '11:30']):
                    header_row = row
                    header_idx = i
                    break
            if not header_row:
                continue
            # Build column index -> canonical name map
            col_map: dict[int, str] = {}
            for j, cell in enumerate(header_row):
                key = str(cell).strip().lower() if cell else ''
                if key in SLOT_COL_MAP:
                    col_map[j] = SLOT_COL_MAP[key]
            slot_cols = list(col_map.items())
            for row in rows[header_idx + 1:]:
                code = normalise_code(row[1] if len(row) > 1 else None)
                if not code:
                    continue
                slots: dict[str, int] = {}
                for col_idx, field in slot_cols:
                    if col_idx < len(row):
                        v = parse_int(row[col_idx])
                        if v is not None and v >= 0:
                            slots[field] = v
                if not slots:
                    continue
                pga = sum(slots.values())
                if pga > 0:
                    add(records, code, iso_date, pga, priority=3)
                    count += 1
    except Exception as e:
        print(f'  [warn] slot {os.path.basename(path)}: {e}')
    return count


# ── PDF: 4-week average sheets ───────────────────────────────────────────────

def read_4week_pdf(records: Records, path: str) -> int:
    """
    Parses PDFs like "WHM 4 Week PGA Average 03052026.pdf".
    Format: "No.Location<d1><d2><d3><d4>Avg" then rows.
    """
    count = 0
    try:
        reader = PdfReader(path)
        text = ''.join(p.extract_text() or '' for p in reader.pages)

        # Extract year from filename (e.g. 03052026 → 2026)
        fname = os.path.basename(path)
        yr_match = re.search(r'(\d{4})\.pdf$', fname)
        year = int(yr_match.group(1)) if yr_match else 2026

        # Find the header line: "No.Location<d1><d2><d3><d4>Avg"
        # Dates look like "12-Apr", "3-May", etc.
        # The text is usually run together without spaces around codes.
        date_pattern = r'(\d{1,2}-[A-Za-z]{3})'
        header_dates = re.findall(date_pattern, text[:300])
        if not header_dates:
            return 0
        iso_dates = []
        for ds in header_dates:
            d = parse_short_date(ds, year)
            if d:
                iso_dates.append(d.strftime('%Y-%m-%d'))
        if not iso_dates:
            return 0

        # Parse data rows: "1WHNLYA 2,5042,5452,3842,5662,500"
        # Extract WH[A-Z]{4} codes (exactly 6 chars) followed by numeric values.
        # Using a letters-only code pattern prevents row-number-suffix noise.
        row_pattern = re.compile(
            r'\d*\s*(WH[A-Z]{4})\s*([\d,\s]+)'
        )
        for m in row_pattern.finditer(text):
            code = normalise_code(m.group(1))
            if not code:
                continue
            vals_raw = re.findall(r'[\d,]+', m.group(2))
            vals = [parse_int(v) for v in vals_raw]
            vals = [v for v in vals if v is not None and v > 0]
            # Expect at least len(iso_dates) values; last value may be the avg
            for i, iso_date in enumerate(iso_dates):
                if i < len(vals):
                    add(records, code, iso_date, vals[i], priority=0)
                    count += 1
    except Exception as e:
        print(f'  [warn] {os.path.basename(path)}: {e}')
    return count


# ── PDF: DIA reports (single week, PGA column) ───────────────────────────────

def read_dia_pdf(records: Records, path: str) -> int:
    """
    Parses WHM DIA PDFs.  Format per row:
      "14 WHKTKT Baalessanvus 11 71 710 332 289"
    where the last number is PGA.
    """
    count = 0
    try:
        reader = PdfReader(path)
        text = ''.join(p.extract_text() or '' for p in reader.pages)

        # Date from first line: "WHM DIA Report - 10.05.2026"
        date_match = re.search(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})', text[:200])
        if not date_match:
            # Try "10th May 2026" from filename
            fname = os.path.basename(path)
            d = parse_date_from_string(fname)
            if not d:
                return 0
            iso_date = d.strftime('%Y-%m-%d')
        else:
            day, mon, yr = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
            iso_date = date(yr, mon, day).strftime('%Y-%m-%d')

        # Each data row ends with: MCA PGA
        # Row pattern: "NN WHCODE <name> <llp_count> <zones> <mcs> <disc> <mca> <pga>"
        row_re = re.compile(
            r'\d+\s+(WH[A-Z0-9]{4,6})\s+\S+.*?(\d+)\s*$',
            re.MULTILINE
        )
        # Simpler: find all WH codes followed by a sequence ending in the PGA value
        # The column order is: LLPs Zones MCs Disciples MCA PGA
        # We just grab: code and then the last number on the logical line
        lines = text.split('\n')
        for line in lines:
            code_m = re.search(r'\b(WH[A-Z]{4})\b', line)
            if not code_m:
                continue
            code = normalise_code(code_m.group(1))
            if not code:
                continue
            # Numbers after the code
            nums = re.findall(r'[\d,]+', line[code_m.end():])
            nums_int = [parse_int(n) for n in nums if parse_int(n)]
            if len(nums_int) >= 2:
                pga = nums_int[-1]  # last number = PGA
                if pga and pga > 0:
                    add(records, code, iso_date, pga, priority=2)
                    count += 1
    except Exception as e:
        print(f'  [warn] {os.path.basename(path)}: {e}')
    return count


# ── Date extraction from individual xlsx filename ────────────────────────────

DATE_IN_FILENAME_RE = re.compile(
    r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)[,\s]+(\d{4})',
    re.IGNORECASE
)


def date_from_pga_filename(path: str) -> str | None:
    fname = os.path.basename(path)
    # Try "14th January, 2023" style
    m = DATE_IN_FILENAME_RE.search(fname)
    if m:
        day, mon, yr = int(m.group(1)), m.group(2).lower()[:3], int(m.group(3))
        if mon in MONTH_MAP:
            return date(yr, MONTH_MAP[mon], day).strftime('%Y-%m-%d')
    # Try "7th July, 2024"
    m2 = re.search(r'(\d{1,2})[A-Za-z]*\s+([A-Za-z]+)\s*[,\s]+(\d{4})', fname, re.IGNORECASE)
    if m2:
        day, mon, yr = int(m2.group(1)), m2.group(2).lower()[:3], int(m2.group(3))
        if mon in MONTH_MAP:
            return date(yr, MONTH_MAP[mon], day).strftime('%Y-%m-%d')
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    records: Records = {}
    data_root = ROOT

    # Load known codes whitelist
    cat_path = ROOT / 'WHM Data' / 'outputs' / 'worship_harvest_categorization_data.json'
    if cat_path.exists():
        with open(cat_path) as f:
            cat = json.load(f)
        for r in cat['records']:
            if r.get('location'):
                KNOWN_CODES.add(r['location'])
        print(f'Loaded {len(KNOWN_CODES)} known location codes')
    else:
        print('[warn] Categorization JSON not found — no whitelist applied')

    print('=== WHM PGA Data Extraction ===')
    print(f'Root: {data_root}')

    # 1. Continuous / 5-week average Excel files (priority 1)
    excel_glob = list(
        glob.glob(str(data_root / 'More Data' / '*.xlsx')) +
        glob.glob(str(data_root / 'Other Data' / '*.xlsx'))
    )
    total_continuous = 0
    for path in sorted(excel_glob):
        n = read_continuous_excel(records, path)
        if n:
            total_continuous += n
    print(f'[Excel continuous/avg] {total_continuous} records from {len(excel_glob)} files')

    # 2. Individual slot Excel files (priority 3)
    slot_candidates = [
        p for p in excel_glob
        if any(kw in os.path.basename(p).lower()
               for kw in ['physical garage attendance', 'pga report', 'pga reports'])
        and 'average' not in os.path.basename(p).lower()
        and 'continuous' not in os.path.basename(p).lower()
    ]
    total_slots = 0
    for path in sorted(slot_candidates):
        iso_date = date_from_pga_filename(path)
        if iso_date:
            n = read_slot_excel(records, path, iso_date)
            total_slots += n
    print(f'[Excel slot sheets]    {total_slots} records from {len(slot_candidates)} files')

    # 3. 4-week average PDFs — WHM Data folder (priority 0)
    pdf_4wk = glob.glob(str(data_root / 'WHM Data' / 'WHM 4 Week PGA*.pdf'))
    total_4wk = 0
    for path in sorted(pdf_4wk):
        n = read_4week_pdf(records, path)
        total_4wk += n
    print(f'[PDF 4-week avg]       {total_4wk} records from {len(pdf_4wk)} files')

    # 4. DIA PDFs — WHM Data folder (priority 2)
    dia_pdfs = glob.glob(str(data_root / 'WHM Data' / 'WHM DIA 2026*.pdf'))
    total_dia = 0
    for path in sorted(dia_pdfs):
        n = read_dia_pdf(records, path)
        total_dia += n
    print(f'[PDF DIA weekly]       {total_dia} records from {len(dia_pdfs)} files')

    # 5. 2025 May PDF
    may25 = glob.glob(str(data_root / 'More Data' / 'WHM May PGA Average 2025.pdf'))
    total_may25 = 0
    for path in may25:
        n = read_4week_pdf(records, path)
        total_may25 += n
    print(f'[PDF 2025 May]         {total_may25} records from {len(may25)} files')

    # Build output list, sorted by date then code
    out_records = []
    for (code, iso_date), meta in sorted(records.items()):
        out_records.append({
            'locationCode': code,
            'weekDate': iso_date,
            'pga': meta['pga'],
        })

    # Summary
    codes = {r['locationCode'] for r in out_records}
    dates = {r['weekDate'] for r in out_records}
    print(f'\nTotal: {len(out_records)} records | {len(codes)} locations | {len(dates)} weeks')
    if dates:
        print(f'Date range: {min(dates)} → {max(dates)}')

    # Warn about gap
    sorted_dates = sorted(dates)
    for i in range(1, len(sorted_dates)):
        prev = datetime.fromisoformat(sorted_dates[i - 1])
        curr = datetime.fromisoformat(sorted_dates[i])
        gap_days = (curr - prev).days
        if gap_days > 45:
            print(f'[GAP] {sorted_dates[i-1]} → {sorted_dates[i]} ({gap_days} days)')

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump({
            'extractedAt': datetime.utcnow().isoformat() + 'Z',
            'recordCount': len(out_records),
            'locationCount': len(codes),
            'weekCount': len(dates),
            'records': out_records,
        }, f, indent=2)

    print(f'\nOutput: {OUTPUT_FILE}')


if __name__ == '__main__':
    main()
